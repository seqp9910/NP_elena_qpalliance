#!/usr/bin/env python3
"""
Elena NP — Backend Flask
Pipeline de Notificaciones Personales para QPAlliance
"""
import os, re, json, uuid, threading, datetime, shutil, subprocess
import unicodedata, base64, zipfile, io, sys, time
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_file, Response

# ─── APP SETUP ───────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB

BASE_DIR   = Path(__file__).parent
ASSETS_DIR = BASE_DIR / 'assets'
JOBS_DIR   = BASE_DIR / 'jobs'
JOBS_DIR.mkdir(exist_ok=True)

JOBS: dict = {}

BREVO_API_KEY = os.environ.get('BREVO_API_KEY', '')
SENDER_NAME   = 'Notificaciones Judiciales - QPAlliance'
SENDER_EMAIL  = 'notificacionesjudiciales@qpalliance.co'

# Load logo base64 once at startup
_LOGO_B64 = ''
try:
    _LOGO_B64 = (ASSETS_DIR / 'logo_b64.txt').read_text().strip()
except Exception:
    pass

# ─── UTILS ───────────────────────────────────────────────────────────────────
def hoy_str():
    m = ['','enero','febrero','marzo','abril','mayo','junio',
         'julio','agosto','septiembre','octubre','noviembre','diciembre']
    t = datetime.date.today()
    return f"{t.day:02d} de {m[t.month]} de {t.year}"

MESES_STR = ['', 'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
             'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
MESES_MAP  = {m: i for i, m in enumerate(MESES_STR) if m}

def fecha_a_letras(fecha_str: str) -> str:
    """Convert date string to Spanish letters: '05 de enero de 2025'."""
    if not fecha_str:
        return fecha_str
    s = str(fecha_str).strip()
    # Already in letter format?
    if re.search(r'\bde\s+(' + '|'.join(MESES_STR[1:]) + r')\b', s, re.I):
        return s
    # dd/mm/yyyy or dd-mm-yyyy
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', s)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= month <= 12:
            return f"{day:02d} de {MESES_STR[month]} de {year}"
    return s

def parse_fecha(s) -> datetime.date | None:
    """Parse a date from string or date/datetime object. Returns datetime.date or None."""
    if isinstance(s, datetime.datetime):
        return s.date()
    if isinstance(s, datetime.date):
        return s
    s = str(s).strip()
    # dd de mes de yyyy
    m = re.match(r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})', s, re.I)
    if m:
        d, mes, y = int(m.group(1)), m.group(2).lower(), int(m.group(3))
        month = MESES_MAP.get(mes)
        if month:
            try:
                return datetime.date(y, month, d)
            except Exception:
                pass
    # dd/mm/yyyy or dd-mm-yyyy
    m = re.match(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', s)
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except Exception:
            pass
    return None

def _pdf_page_to_base64(pdf_path: Path, page_index: int = 0) -> str | None:
    """Render a PDF page to a PNG and return its base64 string."""
    try:
        import fitz  # pymupdf
        doc = fitz.open(str(pdf_path))
        if page_index >= len(doc):
            page_index = 0
        page = doc[page_index]
        mat  = fitz.Matrix(2.0, 2.0)          # 2× zoom → ~144 dpi
        pix  = page.get_pixmap(matrix=mat, alpha=False)
        png_bytes = pix.tobytes('png')
        doc.close()
        import base64
        return base64.standard_b64encode(png_bytes).decode()
    except Exception:
        return None


def extract_fecha_admite_from_pdf(pdf_path: Path, log_fn=None) -> str | None:
    """
    Extract admission date from an auto admisorio PDF.
    Strategy 1 (primary): Claude vision API — reads the page as an image,
                          works with both digital and scanned PDFs.
    Strategy 2 (fallback): pdfplumber text extraction + regex.
    Returns e.g. '05 de enero de 2025' or None.
    """
    def _log(msg):
        if log_fn:
            log_fn(f"    [AA] {msg}")

    MESES_RE = '(' + '|'.join(MESES_STR[1:]) + ')'
    DATE_WORD = r'\b(\d{1,2})\s+de\s+' + MESES_RE + r'\s+de\s+(\d{4})\b'
    DATE_NUM  = r'\b(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})\b'

    def to_letras_from_word(m):
        return f"{int(m.group(1)):02d} de {m.group(2).lower()} de {m.group(3)}"

    def to_letras_from_num(m):
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{d:02d} de {MESES_STR[mo]} de {y}"
        return None

    def parse_date_str(raw: str) -> str | None:
        raw = raw.strip()
        if not raw or raw.upper() == 'NO_FECHA':
            return None
        m = re.search(DATE_WORD, raw, re.IGNORECASE)
        if m:
            return to_letras_from_word(m)
        m = re.search(DATE_NUM, raw)
        if m:
            return to_letras_from_num(m)
        return None

    # ── Strategy 1: Claude vision API ────────────────────────────────────────
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        _log("ANTHROPIC_API_KEY no configurada — usando fallback regex")
    else:
        _log("Intentando extracción con Claude vision...")
        img_b64 = _pdf_page_to_base64(pdf_path, page_index=0)
        if not img_b64:
            _log("No se pudo renderizar la página como imagen (pymupdf falló)")
        else:
            try:
                import anthropic as _anthropic
                client = _anthropic.Anthropic(api_key=api_key)
                msg = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=100,
                    messages=[{
                        'role': 'user',
                        'content': [
                            {
                                'type': 'image',
                                'source': {
                                    'type': 'base64',
                                    'media_type': 'image/png',
                                    'data': img_b64,
                                },
                            },
                            {
                                'type': 'text',
                                'text': (
                                    'Este es un auto admisorio de una demanda laboral colombiana. '
                                    'Extrae la fecha en que fue proferido o admitido el auto. '
                                    'Responde ÚNICAMENTE con la fecha en formato: '
                                    'DD de mes de YYYY (ejemplo: 15 de marzo de 2024). '
                                    'Si no encuentras ninguna fecha clara, responde exactamente: NO_FECHA'
                                ),
                            },
                        ],
                    }],
                )
                raw_answer = msg.content[0].text.strip()
                _log(f"Respuesta IA: '{raw_answer}'")
                result = parse_date_str(raw_answer)
                if result:
                    _log(f"Fecha extraída por IA: {result}")
                    return result
                else:
                    _log("IA no devolvió fecha válida — usando fallback regex")
            except Exception as _e:
                _log(f"Error llamando Claude API: {_e} — usando fallback regex")

    # ── Strategy 2: pdfplumber text + regex (digital PDFs) ───────────────────
    try:
        import pdfplumber
        text = ''
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages[:5]:
                t = page.extract_text()
                if t:
                    text += t + '\n'
    except Exception:
        return None

    if not text.strip():
        return None

    keywords = [
        r'auto\s+(?:de\s+)?fecha',
        r'adiado',
        r'admiti[oó]',
        r'se\s+admite',
        r'auto\s+admisorio',
        r'providencia\s+de\s+fecha',
        r'resoluci[oó]n\s+de\s+fecha',
        r'auto\s+del\s+d[ií]a',
    ]
    for kw in keywords:
        m = re.search(kw + r'(?:[^\n]{0,120}?)' + DATE_WORD, text, re.IGNORECASE | re.DOTALL)
        if m:
            return to_letras_from_word(m)
        m = re.search(kw + r'(?:[^\n]{0,60}?)' + DATE_NUM, text, re.IGNORECASE)
        if m:
            result = to_letras_from_num(m)
            if result:
                return result

    m = re.search(DATE_WORD, text, re.IGNORECASE)
    if m:
        return to_letras_from_word(m)
    m = re.search(DATE_NUM, text)
    if m:
        return to_letras_from_num(m)

    return None

def normalizar(s: str) -> str:
    """Lowercase + strip accents."""
    s = s.strip().lower()
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

def parse_codigos(raw: str) -> list:
    """
    Parse codes from semicolon/comma/space/newline separated string.
    Accepts: '1372;1496'  OR  'R1372;R1496'  OR  'r1372, r1496'
    Returns list of unique ints.
    """
    result = []
    for part in re.split(r'[;,\n\r\t ]+', raw.strip()):
        part = part.strip()
        if not part:
            continue
        m = re.match(r'^[Rr]?(\d+)$', part)
        if m:
            result.append(int(m.group(1)))
    return list(dict.fromkeys(result))  # deduplicate preserving order

_PROMPT_AA = (
    'Este es un auto admisorio de una demanda laboral colombiana.\n'
    'Necesito exactamente dos datos. Responde SOLO en este formato:\n'
    'NOMBRE: [nombre completo de la parte DEMANDANTE]\n'
    'FECHA: [DD de mes de YYYY]\n\n'
    'Instrucciones:\n'
    '- NOMBRE: busca la etiqueta "DEMANDANTE:", "Demandante:", "PARTE DEMANDANTE:", '
    '"Accionante:" o similar. Copia el nombre completo que aparece ahí. '
    'NO copies el nombre del demandado ni del juzgado.\n'
    '- FECHA: fecha en que fue proferido o admitido el auto (no la de radicación).\n'
    '- Si no encuentras un dato escribe NO_ENCONTRADO.\n'
    '- Cero texto adicional fuera del formato.'
)

def _parse_aa_response(raw: str) -> dict:
    """Parse NOMBRE/FECHA lines from Claude response."""
    nombre = ''
    fecha  = ''
    for line in raw.splitlines():
        line = line.strip()
        if line.upper().startswith('NOMBRE:'):
            val = line[7:].strip()
            if val and val.upper() != 'NO_ENCONTRADO':
                nombre = val
        elif line.upper().startswith('FECHA:'):
            val = line[6:].strip()
            if val and val.upper() != 'NO_ENCONTRADO':
                fecha = val
    return {'nombre': nombre, 'fecha': fecha}


def _normalizar_fecha_letras(fecha_str: str) -> str:
    """Normalize any date string to '05 de enero de 2025' format."""
    MESES_RE  = '(' + '|'.join(MESES_STR[1:]) + ')'
    DATE_WORD = r'(\d{1,2})\s+de\s+' + MESES_RE + r'\s+de\s+(\d{4})'
    DATE_NUM  = r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})'
    m = re.search(DATE_WORD, fecha_str, re.IGNORECASE)
    if m:
        return f"{int(m.group(1)):02d} de {m.group(2).lower()} de {m.group(3)}"
    m = re.search(DATE_NUM, fecha_str)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{d:02d} de {MESES_STR[mo]} de {y}"
    return fecha_str


def _extract_ciudad_from_juzgado(juzgado: str) -> str:
    """Extract city from 'Juzgado 14 Laboral del Circuito de Bogotá' → 'Bogotá'."""
    if not juzgado:
        return 'N/A'
    m = re.search(r'\bde\s+([A-ZÁÉÍÓÚÑ][a-záéíóúñA-ZÁÉÍÓÚÑ ]+?)\s*$', juzgado)
    if m:
        return m.group(1).strip()
    return juzgado.strip()


def scan_auto_admisorio(pdf_path: Path, log_fn=None) -> dict:
    """
    Scan an auto admisorio PDF with Claude to extract demandante name + date.

    Strategy A (digital PDFs): extract text with pdfplumber → send as text to
                               Claude (faster, cheaper, no image rendering).
    Strategy B (scanned PDFs): render pages as images → send to Claude vision.

    Tries pages 0 and 1 before giving up.
    Returns {'nombre': str, 'fecha': str} — either may be ''.
    """
    def _log(msg):
        if log_fn:
            log_fn(f"    [scan-AA] {msg}")

    empty = {'nombre': '', 'fecha': ''}
    api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    # Debug: list all env keys starting with ANTH (only first call)
    anth_keys = [k for k in os.environ if 'ANTH' in k.upper()]
    _log(f"Iniciando scan | API key: {'OK ('+str(len(api_key))+' chars)' if api_key else 'FALTA'} | vars ANTH: {anth_keys} | {pdf_path.name}")
    if not api_key:
        _log("ERROR: Sin ANTHROPIC_API_KEY — no se puede escanear")
        return empty

    # ── Strategy A: pdfplumber text extraction ────────────────────────────
    def _extract_text() -> str:
        try:
            import pdfplumber
            text = ''
            with pdfplumber.open(str(pdf_path)) as pdf:
                for page in pdf.pages[:3]:
                    t = page.extract_text() or ''
                    text += t + '\n'
            return text.strip()
        except Exception:
            return ''

    def _call_claude_text(text: str) -> dict:
        try:
            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=api_key)
            prompt = (
                _PROMPT_AA + '\n\nTexto del documento:\n"""\n' +
                text[:4000] + '\n"""'
            )
            msg = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=200,
                messages=[{'role': 'user', 'content': prompt}],
            )
            raw = msg.content[0].text.strip()
            _log(f"Texto → IA: {raw[:120]}")
            return _parse_aa_response(raw)
        except Exception as e:
            _log(f"Error API texto: {e}")
            return empty

    # ── Strategy B: image (vision) ────────────────────────────────────────
    def _call_claude_image(page_idx: int) -> dict:
        img_b64 = _pdf_page_to_base64(pdf_path, page_index=page_idx)
        if not img_b64:
            _log(f"No se pudo renderizar pág {page_idx}")
            return empty
        try:
            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=api_key)
            msg = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=200,
                messages=[{
                    'role': 'user',
                    'content': [
                        {'type': 'image',
                         'source': {'type': 'base64',
                                    'media_type': 'image/png',
                                    'data': img_b64}},
                        {'type': 'text', 'text': _PROMPT_AA},
                    ],
                }],
            )
            raw = msg.content[0].text.strip()
            _log(f"Imagen pág {page_idx} → IA: {raw[:120]}")
            return _parse_aa_response(raw)
        except Exception as e:
            _log(f"Error API imagen pág {page_idx}: {e}")
            return empty

    # Try text first
    text = _extract_text()
    if len(text) > 100:
        _log(f"Texto extraído ({len(text)} chars) → usando Strategy A")
        result = _call_claude_text(text)
    else:
        _log("PDF sin texto suficiente → usando Strategy B (visión)")
        result = _call_claude_image(0)
        if not result.get('nombre'):
            _log("Pág 0 sin nombre, intentando pág 1...")
            r1 = _call_claude_image(1)
            if r1.get('nombre'):
                result['nombre'] = r1['nombre']
            if not result.get('fecha') and r1.get('fecha'):
                result['fecha'] = r1['fecha']

    # Normalize fecha format
    if result.get('fecha'):
        result['fecha'] = _normalizar_fecha_letras(result['fecha'])

    _log(f"  → nombre='{result.get('nombre','')}' | fecha='{result.get('fecha','')}'")
    return result


def _nombre_score(a: str, b: str) -> float:
    """Token overlap similarity between two name strings (0.0–1.0)."""
    ta = set(normalizar(a).split())
    tb = set(normalizar(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / max(len(ta), len(tb))


def build_code_pdf_map(pdf_paths: list, excel_data: dict,
                       valid_codes: list, log_fn=None,
                       doc_type: str = 'demanda',
                       extracted_fechas: dict = None) -> dict:
    """
    Build a {code -> pdf_path} mapping in one pre-processing pass.

    doc_type='auto':
      P1. Filename contains R{code}.
      P2. Claude vision reads demandante name + fecha from the PDF;
          name is matched against Nombre column in excel_data.
          If extracted_fechas dict is provided, it is populated: {code: fecha_str}.
      P3. Positional fallback.

    doc_type='demanda':
      P1. Filename contains R{code}.
      P2. Code digits in filename (unique match only).
      P3. Positional fallback.

    Returns dict[int, Path].
    """
    def _log(msg):
        if log_fn:
            log_fn(f"    [PDF-match] {msg}")

    if not pdf_paths:
        return {}

    # Nombre → code lookup (for auto matching)
    nombre_to_code: dict[str, int] = {}
    for code in valid_codes:
        nombre = excel_data.get(code, {}).get('Nombre', '')
        if nombre:
            nombre_to_code[nombre] = code

    # ── Pre-scan ALL auto PDFs before matching ────────────────────────────
    # This ensures fecha is available regardless of which matching path (P1/P2/P3) wins.
    pdf_scan_cache: dict[str, dict] = {}
    if doc_type == 'auto':
        _log(f"Pre-escaneando {len(pdf_paths)} auto(s) admisorio(s) con IA...")
        for pp in pdf_paths:
            pdf_scan_cache[str(pp)] = scan_auto_admisorio(pp, log_fn=log_fn)

    result: dict[int, Path] = {}
    unmatched_paths: list[Path] = []

    for pdf_path in pdf_paths:
        matched_code = None
        cached = pdf_scan_cache.get(str(pdf_path), {})
        nombre_pdf = cached.get('nombre', '')
        fecha_pdf  = cached.get('fecha', '')

        # ── P1: R{code} in filename ───────────────────────────────────────
        for code in valid_codes:
            if code in result:
                continue
            if re.search(r'[Rr]0*' + str(code) + r'[\W_\.]', pdf_path.name + '.'):
                matched_code = code
                _log(f"{pdf_path.name} → R{code} (nombre archivo)")
                break

        # ── P2 auto: match by demandante name ────────────────────────────
        if matched_code is None and doc_type == 'auto':
            if nombre_pdf:
                best_code  = None
                best_score = 0.0
                for excel_nombre, code in nombre_to_code.items():
                    if code in result:
                        continue
                    score = _nombre_score(nombre_pdf, excel_nombre)
                    if score > best_score:
                        best_score = score
                        best_code  = code

                if best_score >= 0.4:
                    matched_code = best_code
                    _log(f"{pdf_path.name} → R{matched_code} "
                         f"(nombre IA '{nombre_pdf}', score={best_score:.2f})")
                else:
                    _log(f"{pdf_path.name}: score={best_score:.2f} insuficiente "
                         f"(nombre='{nombre_pdf}') → fallback posicional")
            else:
                _log(f"{pdf_path.name}: sin nombre extraído → fallback posicional")

        # ── P2 demanda: code digits in stem (unique) ──────────────────────
        if matched_code is None and doc_type == 'demanda':
            candidates = [
                c for c in valid_codes
                if c not in result and str(c) in re.sub(r'\D', '', pdf_path.stem)
            ]
            if len(candidates) == 1:
                matched_code = candidates[0]
                _log(f"{pdf_path.name} → R{matched_code} (dígitos en nombre)")

        if matched_code is not None:
            result[matched_code] = pdf_path
            # Store fecha for ANY matched code (P1, P2, or P3) if we have it
            if doc_type == 'auto' and extracted_fechas is not None and fecha_pdf:
                extracted_fechas[matched_code] = fecha_pdf
                _log(f"  Fecha guardada R{matched_code}: {fecha_pdf}")
        else:
            unmatched_paths.append(pdf_path)

    # ── P3: positional for unmatched remainder ────────────────────────────
    unmatched_codes = sorted(c for c in valid_codes if c not in result)
    if unmatched_paths and unmatched_codes:
        _log(f"⚠ Posicional: {[p.name for p in unmatched_paths]} → R{unmatched_codes}")
        for code, path in zip(unmatched_codes,
                              sorted(unmatched_paths, key=lambda p: p.name)):
            result[code] = path
            cached_p3 = pdf_scan_cache.get(str(path), {})
            fecha_p3  = cached_p3.get('fecha', '')
            if doc_type == 'auto' and extracted_fechas is not None and fecha_p3:
                extracted_fechas[code] = fecha_p3
                _log(f"  {path.name} → R{code} (posicional) | fecha: {fecha_p3}")
            else:
                _log(f"  {path.name} → R{code} (posicional) | sin fecha")

    return result


def find_pdf_for_code(code: int, paths: list, all_codes: list = None) -> Path | None:
    """Legacy single-lookup (kept for compatibility). Prefer build_code_pdf_map."""
    if not paths:
        return None
    code_str = str(code)
    for p in paths:
        if re.search(r'[Rr]0*' + code_str + r'[\W_\.]', p.name + '.'):
            return p
    for p in paths:
        if code_str in re.sub(r'\D', '', p.stem):
            return p
    if all_codes and len(paths) == len(all_codes):
        sorted_codes = sorted(all_codes)
        sorted_paths = sorted(paths, key=lambda p: p.name)
        if code in sorted_codes:
            return sorted_paths[sorted_codes.index(code)]
    if len(paths) == 1:
        return paths[0]
    return None

def load_excel(excel_path: Path) -> dict:
    """
    Load the NP base Excel file.
    Returns dict: {code_int: {'Nombre', 'Radicado', 'Ciudad', 'Juzgado',
                               'Fecha_Demanda', 'email'}}
    """
    import pandas as pd
    import openpyxl

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo Excel: {e}")

    if 'Total' not in sheets:
        raise ValueError(
            f"El archivo Excel no tiene una hoja llamada 'Total'. "
            f"Hojas encontradas: {sheets}. "
            f"Por favor verifica que el archivo tenga una hoja 'Total'."
        )

    df = pd.read_excel(excel_path, sheet_name='Total', header=0)

    col_map = {}
    cols_list = list(df.columns)
    for idx, col in enumerate(cols_list):
        norm = normalizar(str(col))
        # Column B (index 1) with header '#' is always the case code
        if norm == '#' or (idx == 1 and norm in ('#', 'num', 'numero', 'n', 'no', 'cod', 'codigo', '')):
            col_map['num'] = col
        elif norm in ('num', 'numero') and 'num' not in col_map:
            col_map['num'] = col
        elif 'nombre' in norm and 'nombre' not in col_map:
            col_map['nombre'] = col
        elif 'radicado' in norm and 'radicado' not in col_map:
            col_map['radicado'] = col
        elif 'ciudad' in norm and 'ciudad' not in col_map:
            col_map['ciudad'] = col
        elif 'juzgado' in norm and 'juzgado' not in col_map:
            col_map['juzgado'] = col
        elif 'fecha' in norm and 'demanda' in norm and 'fecha_demanda' not in col_map:
            col_map['fecha_demanda'] = col
        elif ('correo' in norm or 'email' in norm or 'electroni' in norm) and 'email' not in col_map:
            col_map['email'] = col

    result = {}
    num_col = col_map.get('num')
    if num_col is None:
        if '#' in df.columns:
            num_col = '#'
        else:
            raise ValueError(
                f"No se encontro columna de codigo. "
                f"Columnas: {list(df.columns)}"
            )

    for _, row in df.iterrows():
        try:
            raw_code = row[num_col]
            if pd.isna(raw_code):
                continue
            code_str = str(raw_code).strip()
            m = re.match(r'^[Rr]?(\d+)', code_str)
            if not m:
                continue
            code = int(m.group(1))
        except Exception:
            continue

        def get_col(key, default=''):
            col = col_map.get(key)
            if not col:
                return default
            val = row.get(col, default)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return default
            return str(val).strip()

        # Fecha_Demanda: convert to legible string
        fd_col = col_map.get('fecha_demanda')
        fd_raw = row.get(fd_col) if fd_col else None
        if fd_raw is not None and not (isinstance(fd_raw, float) and pd.isna(fd_raw)):
            try:
                if isinstance(fd_raw, (datetime.datetime, datetime.date)):
                    fecha_demanda = fecha_a_letras(fd_raw.strftime('%d/%m/%Y'))
                else:
                    fecha_demanda = fecha_a_letras(str(fd_raw).strip())
            except Exception:
                fecha_demanda = str(fd_raw).strip()
        else:
            fecha_demanda = ''

        juzgado = get_col('juzgado')
        # Ciudad: use explicit column if present, else extract from Juzgado name
        ciudad = get_col('ciudad') or _extract_ciudad_from_juzgado(juzgado)

        result[code] = {
            'Nombre':        get_col('nombre'),
            'Radicado':      get_col('radicado'),
            'Ciudad':        ciudad,
            'Juzgado':       juzgado,
            'Fecha_Demanda': fecha_demanda,
            'email':         get_col('email'),
        }

    return result

def _fix_split_placeholders(xml_text: str, replacements: dict) -> str:
    """
    Replace {FieldName} placeholders in DOCX XML even when split across 3 w:r runs:
      run1: <w:t>{ or ' {'</w:t>
      run2: <w:t>FieldName</w:t>
      run3: <w:t>}</w:t>
    Preserves any leading space before the { in the first run.
    """
    for field, value in replacements.items():
        if str(field).startswith('__'):
            continue  # Skip internal flags
        safe_val = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        # Pass 1: simple single-run replacements
        for variant in ['{'+field+'}', '{ '+field+'}', '{'+field+' }', '{ '+field+' }']:
            xml_text = xml_text.replace(variant, safe_val)

        # Pass 2: split-run — field name appears alone in a w:t element
        field_pat = r'<w:t(?:\s[^>]*)?>' + re.escape(field) + r'</w:t>'
        for m_field in list(re.finditer(field_pat, xml_text)):
            pos = m_field.start()

            before = xml_text[max(0, pos-1500):pos]
            opens = list(re.finditer(r'<w:t(?:\s[^>]*)?>[ ]?\{</w:t>', before))
            if not opens:
                continue
            last_open = opens[-1]
            open_abs_start = max(0, pos-1500) + last_open.start()
            open_abs_end   = max(0, pos-1500) + last_open.end()

            after = xml_text[m_field.end():m_field.end()+500]
            m_close = re.search(r'<w:t(?:\s[^>]*)?>}</w:t>', after)
            if not m_close:
                continue
            close_abs_start = m_field.end() + m_close.start()
            close_abs_end   = m_field.end() + m_close.end()

            # Preserve leading space before { when replacing
            new_open = re.sub(
                r'(<w:t[^>]*>)([ ]?)\{(</w:t>)',
                lambda m, val=safe_val: m.group(1) + m.group(2) + val + m.group(3),
                xml_text[open_abs_start:open_abs_end]
            )
            new_field = re.sub(
                r'(<w:t[^>]*>)' + re.escape(field) + r'(</w:t>)',
                lambda m: m.group(1) + m.group(2),
                xml_text[m_field.start():m_field.end()]
            )
            new_close = re.sub(
                r'(<w:t[^>]*>)}(</w:t>)',
                lambda m: m.group(1) + m.group(2),
                xml_text[close_abs_start:close_abs_end]
            )

            xml_text = (xml_text[:open_abs_start] + new_open +
                        xml_text[open_abs_end:m_field.start()] + new_field +
                        xml_text[m_field.end():close_abs_start] + new_close +
                        xml_text[close_abs_end:])
            break
    return xml_text


def fill_template(template_path: Path, data: dict, output_path: Path,
                  no_date_mode: bool = False):
    """
    Replace {placeholder} in DOCX template (handles split-run XML) and save.
    If no_date_mode=True, replaces the 'adiado {fecha_admite}' segment with
    '-y que aquí se adjunta-' (alternate paragraph when date not found).
    """
    import shutil as _shutil
    import zipfile as _zipfile

    _shutil.copy2(template_path, output_path)

    with _zipfile.ZipFile(output_path, 'r') as z:
        names = z.namelist()
        contents = {n: z.read(n) for n in names}

    all_data = dict(data)
    all_data['fecha_de_hoy'] = hoy_str()
    if no_date_mode:
        all_data['fecha_admite'] = '__NOFECHA__'

    def replace_in_xml(xml_bytes: bytes) -> bytes:
        text = xml_bytes.decode('utf-8')
        text = _fix_split_placeholders(text, all_data)
        if no_date_mode:
            # Remove ', adiado ' from the run that precedes {fecha_admite}
            text = text.replace('demanda, adiado </w:t>', 'demanda</w:t>')
            text = text.replace('demanda, adiado</w:t>', 'demanda</w:t>')
            # Replace the dummy token with alternate clause
            text = text.replace('__NOFECHA__', '-y que aqu\u00ed se adjunta-')
        return text.encode('utf-8')

    xml_files = [n for n in names if n.endswith('.xml') or n.endswith('.rels')]

    with _zipfile.ZipFile(output_path, 'w', _zipfile.ZIP_DEFLATED) as zout:
        for name in names:
            if name in xml_files:
                zout.writestr(name, replace_in_xml(contents[name]))
            else:
                zout.writestr(name, contents[name])

def docx_to_pdf(docx_path: Path, output_dir: Path) -> Path:
    """Convert DOCX to PDF using LibreOffice headless."""
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'pdf',
         '--outdir', str(output_dir), str(docx_path)],
        capture_output=True, text=True, timeout=60
    )
    pdf_name = docx_path.stem + '.pdf'
    pdf_path = output_dir / pdf_name
    if not pdf_path.exists():
        raise RuntimeError(
            f"LibreOffice fallo convirtiendo {docx_path.name}: {result.stderr}"
        )
    return pdf_path

def merge_pdfs(pdf_list: list, output_path: Path):
    """Merge list of PDF paths into a single PDF using pypdf."""
    from pypdf import PdfWriter
    writer = PdfWriter()
    for p in pdf_list:
        writer.append(str(p))
    with open(output_path, 'wb') as f:
        writer.write(f)

def build_separator_page(output_path: Path, text: str = 'DEMANDA'):
    """
    Build a branded separator page: QPAlliance logo + title text (centered group)
    + footer with address and website.
    """
    import base64, io
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.utils import ImageReader

    # ── Font ─────────────────────────────────────────────────────────────────
    CALADEA = '/usr/share/fonts/truetype/crosextra/Caladea-Bold.ttf'
    font_name = 'CalaDeaBold'
    try:
        pdfmetrics.registerFont(TTFont(font_name, CALADEA))
    except Exception:
        font_name = 'Helvetica-Bold'

    # QPAlliance pink (for footer line only)
    QP_R, QP_G, QP_B = 0.831, 0.0, 0.416   # #D4006A

    w, h = letter
    c = canvas.Canvas(str(output_path), pagesize=letter)

    # ── Auto-size title ───────────────────────────────────────────────────────
    font_size = 60
    c.setFont(font_name, font_size)
    while font_size > 18 and c.stringWidth(text, font_name, font_size) > w - 80:
        font_size -= 4
        c.setFont(font_name, font_size)
    text_w = c.stringWidth(text, font_name, font_size)

    # ── Logo ─────────────────────────────────────────────────────────────────
    logo_img  = None
    logo_w_pt = 0
    logo_h_pt = 0
    if _LOGO_B64:
        try:
            img_bytes  = base64.b64decode(_LOGO_B64)
            logo_img   = ImageReader(io.BytesIO(img_bytes))
            orig_w, orig_h = logo_img.getSize()
            max_logo_w = 160.0
            scale      = min(max_logo_w / orig_w, 80.0 / orig_h)
            logo_w_pt  = orig_w * scale
            logo_h_pt  = orig_h * scale
        except Exception:
            logo_img = None

    # ── Vertical centering: title (top) → gap → logo (bottom) ────────────────
    logo_gap = 24   # pts between title baseline and logo top
    group_h  = font_size + (logo_gap + logo_h_pt if logo_img else 0)
    center_y = h / 2 + 10   # slight upward bias

    # Title baseline sits at the TOP of the group
    title_y  = center_y + group_h / 2 - font_size
    # Logo sits BELOW the title
    logo_y   = center_y - group_h / 2          # bottom of logo image
    logo_x   = (w - logo_w_pt) / 2

    # ── Draw title (BLACK) ────────────────────────────────────────────────────
    c.setFont(font_name, font_size)
    c.setFillColorRGB(0, 0, 0)                  # black
    c.drawString((w - text_w) / 2, title_y, text)

    # ── Draw logo below title ─────────────────────────────────────────────────
    if logo_img:
        c.drawImage(logo_img, logo_x, logo_y,
                    width=logo_w_pt, height=logo_h_pt, mask='auto')

    # ── Footer ────────────────────────────────────────────────────────────────
    FOOT_FONT  = 'Helvetica'
    FOOT_SIZE  = 7.5
    MARGIN     = 36   # left/right margin in pts

    addr_left  = 'Av. Calle 26 #68C-61, Oficinas 909 y 910  \xb7  Edificio Torre Central  \xb7  Bogot\xe1, Colombia'
    addr_right = 'www.qpalliance.co'

    # Full-width pink line — edge to edge
    c.setStrokeColorRGB(QP_R, QP_G, QP_B)
    c.setLineWidth(1.0)
    c.line(0, 46, w, 46)

    # Footer text: left-aligned address, right-aligned URL
    c.setFont(FOOT_FONT, FOOT_SIZE)
    c.setFillColorRGB(0.45, 0.45, 0.45)
    right_w = c.stringWidth(addr_right, FOOT_FONT, FOOT_SIZE)
    c.drawString(MARGIN, 30, addr_left)
    c.drawString(w - MARGIN - right_w, 30, addr_right)

    c.save()


def build_email_proof_pdf(output_path: Path, code: int, client: dict,
                          to_email: str, sent_ok: bool, sent_msg: str,
                          email_subject: str, email_body_text: str):
    """Build a PDF page showing the sent email details as proof of dispatch."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Heading1'],
                                 fontSize=14, textColor=colors.HexColor('#D4006A'))
    label_style = ParagraphStyle('Label', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.grey)
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=10)
    body_style  = ParagraphStyle('Body', parent=styles['Normal'],
                                 fontSize=9, leading=13)

    doc = SimpleDocTemplate(str(output_path), pagesize=letter,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    elements = []

    status_color = '#006600' if sent_ok else '#CC0000'
    status_text  = 'ENVIADO ✓' if sent_ok else 'ERROR AL ENVIAR'

    elements.append(Paragraph('Constancia de Envío — Notificación Personal', title_style))
    elements.append(Spacer(1, 6))
    elements.append(HRFlowable(width='100%', thickness=1,
                                color=colors.HexColor('#D4006A')))
    elements.append(Spacer(1, 10))

    fields = [
        ('Estado',       f'<font color="{status_color}"><b>{status_text}</b></font>'),
        ('Fecha y hora', datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ('De',           SENDER_EMAIL),
        ('Para',         to_email),
        ('Asunto',       email_subject),
        ('Código',       f'R{code}'),
        ('Nombre',       client.get('Nombre', '')),
        ('Radicado',     client.get('Radicado', '')),
    ]
    if not sent_ok:
        fields.append(('Detalle error', sent_msg[:200]))

    for label, value in fields:
        elements.append(Paragraph(label, label_style))
        elements.append(Paragraph(str(value), value_style))
        elements.append(Spacer(1, 4))

    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width='100%', thickness=0.5, color=colors.grey))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph('Cuerpo del correo:', label_style))
    elements.append(Spacer(1, 4))
    plain_body = re.sub(r'<[^>]+>', '', email_body_text).strip()
    for line in plain_body.split('\n'):
        line = line.strip()
        if line:
            elements.append(Paragraph(line, body_style))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f'Generado por Elena NP — QPAlliance — {hoy_str()}',
        ParagraphStyle('Footer', parent=styles['Normal'],
                       fontSize=8, textColor=colors.grey)))
    doc.build(elements)


def build_email_signature() -> str:
    """Build HTML email signature with QPAlliance logo."""
    logo_img = (
        f'<img src="data:image/png;base64,{_LOGO_B64}" '
        f'style="width:90px;height:auto;display:block;" alt="QPAlliance">'
    ) if _LOGO_B64 else '<strong style="color:#D4006A;font-size:16px">qpa</strong>'

    return f"""<br><br>
<hr style="border:none;border-top:2px solid #D4006A;margin:20px 0;max-width:420px">
<table cellpadding="0" cellspacing="0" style="font-family:Arial,sans-serif;font-size:13px;color:#333">
  <tr>
    <td style="padding-right:14px;vertical-align:middle">{logo_img}</td>
    <td style="vertical-align:middle;padding-left:14px;border-left:3px solid #D4006A;line-height:1.6">
      <strong style="font-size:14px;color:#222">Legal Department | QPAlliance</strong><br>
      <span style="color:#666">www.qpalliance.co</span><br>
      <span style="color:#666">notificacionesjudiciales@qpalliance.co</span>
    </td>
  </tr>
</table>"""


def send_email_brevo(to_email: str, to_name: str, subject: str,
                     html_body: str, attachment_paths: list = None):
    """Send email via Brevo transactional API with multiple attachments and signature."""
    import urllib.request
    if not BREVO_API_KEY:
        return False, "BREVO_API_KEY no configurada"

    full_body = html_body + build_email_signature()

    payload = {
        "sender":  {"name": SENDER_NAME, "email": SENDER_EMAIL},
        "to":      [{"email": to_email, "name": to_name}],
        "subject": subject,
        "htmlContent": full_body,
    }

    attachments = []
    for att_path in (attachment_paths or []):
        if att_path is None:
            continue
        p = Path(att_path)
        if p.exists() and p.stat().st_size > 0:
            with open(p, 'rb') as f:
                content_b64 = base64.b64encode(f.read()).decode()
            attachments.append({"content": content_b64, "name": p.name})
    if attachments:
        payload["attachment"] = attachments

    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(
        'https://api.brevo.com/v3/smtp/email',
        data=data,
        headers={
            'Content-Type': 'application/json',
            'api-key': BREVO_API_KEY,
        },
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return True, resp.read().decode()
    except Exception as e:
        return False, str(e)


def build_output_excel(cases_info: list, output_path: Path):
    """Build Excel output with processing metrics per case."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lote NP"

    headers = [
        'Código', 'Nombre', 'Radicado', 'Ciudad',
        'Fecha Radicación Demanda', 'Fecha Admite', 'Resta (días)',
        'Procesó AA', 'Enviado'
    ]

    pink = PatternFill("solid", fgColor="D4006A")
    white_bold = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = pink
        cell.font = white_bold
        cell.alignment = center

    for i, case in enumerate(cases_info, 2):
        fecha_dem = case.get('Fecha_Demanda', '')
        fecha_adm = case.get('fecha_admite_extracted', '')

        resta = ''
        if fecha_dem and fecha_adm:
            try:
                d1 = parse_fecha(fecha_dem)
                d2 = parse_fecha(fecha_adm)
                if d1 and d2:
                    resta = abs((d2 - d1).days)
            except Exception:
                resta = ''

        row_data = [
            f"R{case.get('code', '')}",
            case.get('Nombre', ''),
            case.get('Radicado', ''),
            case.get('Ciudad', ''),
            str(fecha_dem),
            str(fecha_adm) if fecha_adm else '',
            resta,
            '✓' if case.get('procesó_aa') else '✗',
            '✓' if case.get('enviado') else '✗',
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.alignment = Alignment(horizontal='center' if col in (1, 7, 8, 9) else 'left')
            if col == 8:
                cell.font = Font(color='006600' if case.get('procesó_aa') else 'CC0000', bold=True)
            elif col == 9:
                cell.font = Font(color='006600' if case.get('enviado') else 'CC0000', bold=True)

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    wb.save(str(output_path))


def build_receipt_pdf(codes_data: list, output_path: Path):
    """Build a receipt PDF listing all processed cases."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(output_path), pagesize=letter)
    elements = []

    elements.append(Paragraph("Constancia de Notificaciones Personales", styles['Title']))
    elements.append(Paragraph(f"Fecha: {hoy_str()}", styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [['Codigo', 'Nombre', 'Radicado', 'Ciudad', 'Juzgado']]
    for row in codes_data:
        data.append([
            str(row.get('code', '')),
            row.get('Nombre', ''),
            row.get('Radicado', ''),
            row.get('Ciudad', ''),
            row.get('Juzgado', ''),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C4006A')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f0f5')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)

# ─── JOB RUNNER ──────────────────────────────────────────────────────────────
def run_job(job_id: str, job_dir: Path, codigos: list,
            excel_path: Path, autos_pdfs: list, demandas_pdfs: list,
            dest_email: str):
    """Background thread: full NP pipeline for one batch."""
    job = JOBS[job_id]

    def log(msg: str, step: int = None):
        job['log'].append(msg)
        if step is not None:
            job['step'] = step
        print(f"[{job_id}] {msg}")

    try:
        job['status'] = 'running'
        template_path = ASSETS_DIR / 'modelo_NP.docx'

        # STEP 1 — Load Excel
        log("Cargando base de datos Excel...", step=1)
        excel_data = load_excel(excel_path)
        sample_keys = list(excel_data.keys())[:5]
        log(f"OK: {len(excel_data)} registros cargados. Primeros codigos: {sample_keys}")
        log(f"Codigos solicitados: {codigos}")

        found    = [c for c in codigos if c in excel_data]
        missing  = [c for c in codigos if c not in excel_data]
        log(f"Encontrados: {found} | No encontrados: {missing}")
        if not found:
            log("ERROR: Ninguno de los codigos existe en el Excel.")
            job['status'] = 'error'
            job['error']  = f"Codigos {codigos} no encontrados. Primeros en Excel: {sample_keys}"
            return

        # STEP 2 — Build PDF→code mappings and validate
        log("Identificando y validando documentos...", step=2)
        log(f"  Autos subidos: {[p.name for p in autos_pdfs]}")
        log(f"  Demandas subidas: {[p.name for p in demandas_pdfs]}")

        extracted_fechas: dict = {}   # populated by scan_auto_admisorio via build_code_pdf_map
        auto_map    = build_code_pdf_map(autos_pdfs,    excel_data, found, log_fn=log,
                                         doc_type='auto', extracted_fechas=extracted_fechas)
        demanda_map = build_code_pdf_map(demandas_pdfs, excel_data, found, log_fn=log,
                                         doc_type='demanda')

        valid_codes = []
        for code in found:
            missing_docs = []
            if code not in auto_map:
                missing_docs.append('auto admisorio')
            if code not in demanda_map:
                missing_docs.append('demanda')
            if missing_docs:
                log(f"  ERROR R{code}: Faltan documentos requeridos: {', '.join(missing_docs)}. Omitiendo.")
            else:
                log(f"  R{code}: auto={auto_map[code].name} | demanda={demanda_map[code].name}")
                valid_codes.append(code)

        if not valid_codes:
            job['status'] = 'error'
            job['error']  = "Ningún código tiene todos los documentos requeridos (auto admisorio + demanda)."
            log("ERROR: No hay codigos validos con todos los documentos.")
            return

        # STEP 3 — Generate NP, send email, build proof, assemble final PDF per client
        log("Generando notificaciones personales...", step=3)
        cases_info = []
        paquetes   = []

        for code in valid_codes:
            row          = excel_data[code]
            nombre       = row.get('Nombre', '')
            radicado     = row.get('Radicado', '')
            ciudad       = row.get('Ciudad', '')
            juzgado      = row.get('Juzgado', '')
            fecha_demanda = row.get('Fecha_Demanda', '')

            # Locate uploaded PDFs using pre-built mapping (guaranteed to exist)
            auto_pdf    = auto_map[code]
            demanda_pdf = demanda_map[code]

            # 3a. Fecha admite: already extracted during PDF matching (scan_auto_admisorio)
            fecha_admite_extracted = extracted_fechas.get(code) or ''
            no_date_mode = not bool(fecha_admite_extracted)
            if fecha_admite_extracted:
                log(f"  Fecha admisión R{code}: {fecha_admite_extracted}")
            else:
                log(f"  Sin fecha admisión para R{code} — modo sin fecha")

            procesó_aa = bool(fecha_admite_extracted)

            # 3b. Fill NP template → DOCX → PDF
            fill_data = {
                'Nombre':       nombre,
                'Radicado':     radicado,
                'Ciudad':       ciudad,
                'Juzgado':      juzgado,
                'fecha_admite': fecha_admite_extracted or '',
            }
            docx_out = job_dir / f"auto_{code}.docx"
            np_pdf   = None
            try:
                fill_template(template_path, fill_data, docx_out,
                              no_date_mode=no_date_mode)
                np_pdf = docx_to_pdf(docx_out, job_dir)
                log(f"  OK NP generada: R{code} - {nombre}")
            except Exception as e:
                log(f"  Error generando NP R{code}: {e}")

            # 3c. Build legal email body (Quicksand 14pt, two versions)
            email_subject = f"R{code} Notificación personal - {radicado} - {nombre}"
            QS = "font-family:'Quicksand',Arial,sans-serif;font-size:14px;color:#222;line-height:1.7"
            if fecha_admite_extracted:
                email_body = (
                    f'<div style="{QS}">'
                    f'<p>Señores<br>Rappi S.A.S.<br>Felipe Villamarín Lafaurie</p>'
                    f'<p><strong>RADICADO</strong>: {radicado}<br>'
                    f'<strong>REFERENCIA:</strong> Demanda ordinaria laboral promovida por '
                    f'<strong>{nombre}</strong> en contra de Rappi S.A.S.<br>'
                    f'<strong>ASUNTO</strong>: Notificación personal de auto admisorio de '
                    f'demanda ordinaria laboral de primera instancia</p>'
                    f'<p>Reciban un cordial saludo.</p>'
                    f'<p>De manera atenta, y en cumplimiento de lo dispuesto en el artículo 8 '
                    f'de la Ley 2213 de 2022, nos permitimos notificarles el auto de fecha '
                    f'<strong>{fecha_admite_extracted}</strong>, mediante el cual el '
                    f'<strong>{juzgado}</strong> admitió la demanda ordinaria laboral presentada '
                    f'por nuestro representado, el señor {nombre}, en contra de Rappi S.A.S.</p>'
                    f'<p>Para los efectos legales correspondientes, junto con la presente '
                    f'comunicación se remiten los documentos que hacen parte de la actuación '
                    f'procesal y que permiten conocer integralmente el contenido de la providencia '
                    f'y de la demanda presentada, incluyendo el auto admisorio, el escrito de '
                    f'demanda con sus respectivos anexos y pruebas, el poder debidamente otorgado, '
                    f'así como los certificados de existencia y representación legal de las partes '
                    f'y el proyecto de liquidación de pretensiones elaborado para efectos '
                    f'ilustrativos del proceso.</p>'
                    f'<p>La presente notificación se realiza por este medio electrónico en los '
                    f'términos previstos en la normativa vigente, con el fin de garantizar el '
                    f'conocimiento oportuno de la providencia judicial y de la documentación que '
                    f'integra la actuación.</p>'
                    f'<p>Cordialmente,</p>'
                    f'</div>'
                )
            else:
                email_body = (
                    f'<div style="{QS}">'
                    f'<p>Señores<br>Rappi S.A.S.<br>Felipe Villamarín Lafaurie</p>'
                    f'<p><strong>RADICADO</strong>: {radicado}<br>'
                    f'<strong>REFERENCIA:</strong> Demanda ordinaria laboral promovida por '
                    f'<strong>{nombre}</strong> en contra de Rappi S.A.S.<br>'
                    f'<strong>ASUNTO</strong>: Notificación personal de auto admisorio de '
                    f'demanda ordinaria laboral de primera instancia</p>'
                    f'<p>Reciban un cordial saludo.</p>'
                    f'<p>De manera atenta, y en cumplimiento de lo dispuesto en el artículo 8 '
                    f'de la Ley 2213 de 2022, nos permitimos notificarles el auto por medio del '
                    f'cual el <strong>{juzgado}</strong> admitió la demanda ordinaria laboral '
                    f'presentada por nuestro representado, el señor {nombre}, en contra de '
                    f'Rappi S.A.S.</p>'
                    f'<p>Para los efectos legales correspondientes, junto con la presente '
                    f'comunicación se remiten los documentos que hacen parte de la actuación '
                    f'procesal y que permiten conocer integralmente el contenido de la providencia '
                    f'y de la demanda presentada, incluyendo el auto admisorio, el escrito de '
                    f'demanda con sus respectivos anexos y pruebas, el poder debidamente otorgado, '
                    f'así como los certificados de existencia y representación legal de las partes '
                    f'y el proyecto de liquidación de pretensiones elaborado para efectos '
                    f'ilustrativos del proceso.</p>'
                    f'<p>La presente notificación se realiza por este medio electrónico en los '
                    f'términos previstos en la normativa vigente, con el fin de garantizar el '
                    f'conocimiento oportuno de la providencia judicial y de la documentación que '
                    f'integra la actuación.</p>'
                    f'<p>Cordialmente,</p>'
                    f'</div>'
                )

            # 3d. Build all 4 separator pages
            def _sep(name):
                p = job_dir / f"sep_{code}_{name}.pdf"
                try:
                    build_separator_page(p, name)
                    return p
                except Exception as e:
                    log(f"  Error separador '{name}' R{code}: {e}")
                    return None

            sep_np    = _sep('NOTIFICACIÓN PERSONAL')
            sep_aa    = _sep('AUTO ADMISORIO')
            sep_comp  = _sep('Comprobante de notificación personal')
            sep_dem   = _sep('DEMANDA')

            # 3e. Send email FIRST (before proof) — single merged PDF attachment
            # Build pre-proof assembly: NP_sep + NP + AA_sep + auto + DEM_sep + demanda
            pre_proof_parts = [p for p in [sep_np, np_pdf, sep_aa, auto_pdf, sep_dem, demanda_pdf]
                               if p is not None and Path(p).exists()]
            email_att_pdf = None
            if pre_proof_parts:
                email_att_pdf = job_dir / f"R{code}_email.pdf"
                try:
                    merge_pdfs(pre_proof_parts, email_att_pdf)
                except Exception as e:
                    log(f"  Error construyendo adjunto email R{code}: {e}")
                    email_att_pdf = None

            sent_ok  = False
            sent_msg = 'BREVO_API_KEY no configurada'
            if dest_email and BREVO_API_KEY:
                log(f"  Enviando correo R{code} → {dest_email}...")
                sent_ok, sent_msg = send_email_brevo(
                    dest_email, dest_email, email_subject, email_body,
                    [email_att_pdf] if email_att_pdf else [])
                log(f"  {'OK correo enviado' if sent_ok else 'Error correo'} R{code}: {sent_msg[:80]}")
            else:
                log(f"  Email no enviado R{code} (BREVO_API_KEY no configurada)")

            # 3f. Build constancia de envío (proof page)
            proof_pdf = job_dir / f"proof_{code}.pdf"
            try:
                build_email_proof_pdf(proof_pdf, code, row,
                                      dest_email or '(sin destinatario)',
                                      sent_ok, sent_msg, email_subject, email_body)
            except Exception as e:
                log(f"  Error constancia correo R{code}: {e}")
                proof_pdf = None

            # 3g. Merge final output PDF:
            #   NP_sep → NP → AA_sep → auto → Comprobante_sep → proof → DEM_sep → demanda
            final_parts = [p for p in [
                sep_np, np_pdf,
                sep_aa, auto_pdf,
                sep_comp, proof_pdf,
                sep_dem, demanda_pdf,
            ] if p is not None and Path(p).exists()]

            if final_parts:
                paquete_path = job_dir / f"R{code}_DDD_NP_done_pdf"
                try:
                    merge_pdfs(final_parts, paquete_path)
                    paquetes.append(paquete_path)
                    log(f"  OK R{code}_DDD_NP_done_pdf ({len(final_parts)} secciones)")
                except Exception as e:
                    log(f"  Error ensamblando R{code}: {e}")
            else:
                log(f"  Sin partes para R{code}, omitiendo.")

            cases_info.append({
                'code':               code,
                'Nombre':             nombre,
                'Radicado':           radicado,
                'Ciudad':             ciudad,
                'Juzgado':            juzgado,
                'Fecha_Demanda':      fecha_demanda,
                'fecha_admite_extracted': fecha_admite_extracted or '',
                'procesó_aa':         procesó_aa,
                'enviado':            sent_ok,
            })

        # STEP 4 — Build output Excel + constancia
        log("Generando Excel de métricas...", step=4)
        excel_out = job_dir / f'Reporte_NP_{job_id[:8]}.xlsx'
        try:
            build_output_excel(cases_info, excel_out)
            log(f"  OK Excel generado: {excel_out.name}")
        except Exception as e:
            log(f"  Error generando Excel: {e}")
            excel_out = None

        receipt_path = job_dir / 'constancia_NP.pdf'
        try:
            build_receipt_pdf(cases_info, receipt_path)
            log("  OK constancia generada.")
        except Exception as e:
            log(f"  Error generando constancia: {e}")
            receipt_path = None

        # STEP 5 — ZIP
        log("Empaquetando archivos finales...", step=5)
        zip_path = job_dir / f'NP_lote_{job_id[:8]}.zip'
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for p in paquetes:
                zf.write(p, p.name)
            if receipt_path and receipt_path.exists():
                zf.write(receipt_path, receipt_path.name)
            if excel_out and excel_out.exists():
                zf.write(excel_out, excel_out.name)
        log(f"  OK ZIP: {zip_path.name} ({zip_path.stat().st_size // 1024} KB)")

        job['status']    = 'done'
        job['zip_path']  = str(zip_path)
        job['paquetes']  = len(paquetes)
        job['total']     = len(codigos)
        job['cases']     = cases_info
        log("Pipeline completado.", step=6)

    except Exception as e:
        import traceback
        job['status'] = 'error'
        job['error']  = str(e)
        log(f"Error fatal: {e}\n{traceback.format_exc()}")

# ─── ROUTES ──────────────────────────────────────────────────────────────────
@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    raw_codigos = request.form.get('codigos', '').strip()
    if not raw_codigos:
        return jsonify({'error': 'No se ingresaron codigos.'}), 400

    codigos = parse_codigos(raw_codigos)
    if not codigos:
        return jsonify({
            'error': 'No se ingresaron codigos validos. '
                     'Ingrese numeros separados por ";" (ej: 1372;1496)'
        }), 400

    job_id  = str(uuid.uuid4())
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True)

    excel_file = request.files.get('excel') or request.files.get('base_file')
    if not excel_file or not excel_file.filename:
        return jsonify({'error': 'Debe subir el archivo Excel de base.'}), 400

    excel_path = job_dir / 'base.xlsx'
    excel_file.save(str(excel_path))

    autos_files = request.files.getlist('autos_pdfs') or request.files.getlist('autos')
    autos_pdfs  = []
    autos_dir   = job_dir / 'autos'
    autos_dir.mkdir()
    for f in autos_files:
        if f and f.filename:
            p = autos_dir / f.filename
            f.save(str(p))
            autos_pdfs.append(p)

    demandas_files = request.files.getlist('demandas_pdfs') or request.files.getlist('demandas')
    demandas_pdfs  = []
    demandas_dir   = job_dir / 'demandas'
    demandas_dir.mkdir()
    for f in demandas_files:
        if f and f.filename:
            p = demandas_dir / f.filename
            f.save(str(p))
            demandas_pdfs.append(p)

    dest_email = (request.form.get('email') or request.form.get('email_to') or '').strip()

    JOBS[job_id] = {
        'status':  'queued',
        'step':    0,
        'log':     [],
        'codigos': codigos,
    }

    t = threading.Thread(
        target=run_job,
        args=(job_id, job_dir, codigos, excel_path,
              autos_pdfs, demandas_pdfs, dest_email),
        daemon=True
    )
    t.start()

    return jsonify({'job_id': job_id})

@app.route('/status/<job_id>')
def status(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({'error': 'Job no encontrado'}), 404
    total    = job.get('total', len(job.get('codigos', [])))
    step     = job.get('step', 0)
    paquetes = job.get('paquetes', 0)
    progress = int((step / 6) * 100) if step else 0
    return jsonify({
        'status':    job['status'],
        'step':      step,
        'progress':  progress,
        'messages':  job.get('log', []),
        'log':       job.get('log', []),
        'paquetes':  paquetes,
        'sent':      paquetes,
        'total':     total,
        'error':     job.get('error', ''),
        'dash_path': job['status'] == 'done',
    })

@app.route('/download/<job_id>')
def download(job_id):
    job = JOBS.get(job_id)
    if not job or job['status'] != 'done':
        return jsonify({'error': 'Job no listo'}), 404
    zip_path = Path(job['zip_path'])
    if not zip_path.exists():
        return jsonify({'error': 'Archivo no encontrado'}), 404
    return send_file(str(zip_path), as_attachment=True,
                     download_name=zip_path.name)

@app.route('/debug-excel', methods=['GET', 'POST'])
def debug_excel():
    if request.method == 'GET':
        return '''<!DOCTYPE html><html><head><title>Debug Excel</title>
        <style>body{font-family:sans-serif;max-width:600px;margin:80px auto;background:#0a0a0a;color:#fff}
        input,button{margin:10px 0;padding:10px;font-size:16px}button{background:#D4006A;color:#fff;border:none;cursor:pointer;border-radius:6px}
        pre{background:#1a1a1a;padding:15px;border-radius:8px;overflow-x:auto;font-size:13px;max-height:500px;overflow-y:auto}</style></head>
        <body><h2>Debug Excel — Elena NP</h2>
        <form method="POST" enctype="multipart/form-data">
        <p>Sube tu Excel de base de radicación:</p>
        <input type="file" name="base_file" accept=".xlsx,.xls" required><br>
        <button type="submit">Analizar</button></form></body></html>'''
    f = request.files.get('excel') or request.files.get('base_file')
    if not f:
        return jsonify({'error': 'Sube el Excel con campo "excel"'}), 400
    import tempfile, pandas as pd, openpyxl
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = Path(tmp.name)
    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        sheet_name = 'Total' if 'Total' in sheets else sheets[0]
        df = pd.read_excel(tmp_path, sheet_name=sheet_name, header=0)
        cols = list(df.columns)
        num_col = '#' if '#' in df.columns else cols[0]
        sample = [str(row[num_col]) for _, row in df.head(10).iterrows()]
        return jsonify({
            'sheets': sheets, 'sheet_used': sheet_name,
            'columns': [str(c) for c in cols],
            'num_col_detected': num_col,
            'sample_codes_from_num_col': sample,
            'total_rows': len(df),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        tmp_path.unlink(missing_ok=True)

@app.route('/dashboard/<job_id>')
def dashboard(job_id):
    job = JOBS.get(job_id)
    if not job:
        return "Job no encontrado", 404

    cases    = job.get('cases', [])
    paquetes = job.get('paquetes', 0)
    total    = job.get('total', 0)
    errores  = total - paquetes

    # KPI extras
    aa_ok_count   = sum(1 for c in cases if c.get('procesó_aa', False))
    sent_ok_count = sum(1 for c in cases if c.get('enviado', False))

    # Chart 1: NPs by city
    ciudad_counts = {}
    # Chart 2: avg days per city
    ciudad_days = {}
    all_days = []

    for c in cases:
        ciudad = c.get('Ciudad', 'N/A') or 'N/A'
        ciudad_counts[ciudad] = ciudad_counts.get(ciudad, 0) + 1
        fd = c.get('Fecha_Demanda', '')
        fa = c.get('fecha_admite_extracted', '')
        if fd and fa:
            d1 = parse_fecha(fd)
            d2 = parse_fecha(fa)
            if d1 and d2:
                days = abs((d2 - d1).days)
                all_days.append(days)
                ciudad_days.setdefault(ciudad, []).append(days)

    ciudad_labels_json = json.dumps(list(ciudad_counts.keys()))
    ciudad_values_json = json.dumps(list(ciudad_counts.values()))

    avg_cities = list(ciudad_days.keys())
    avg_vals   = [round(sum(v) / len(v), 1) for v in ciudad_days.values()]
    avg_labels_json = json.dumps(avg_cities)
    avg_values_json = json.dumps(avg_vals)

    prom_dias = round(sum(all_days) / len(all_days), 1) if all_days else 'N/A'

    # Table rows
    table_rows = ''
    for c in cases:
        code    = c.get('code', '')
        nombre  = c.get('Nombre', '') or ''
        juzgado = c.get('Juzgado', '') or ''
        radicado= c.get('Radicado', '') or ''
        fd      = c.get('Fecha_Demanda', '') or ''
        fa      = c.get('fecha_admite_extracted', '') or ''
        aa_ok   = c.get('procesó_aa', False)
        flujo_ok= c.get('enviado', False)
        tramite_str = '—'
        if fd and fa:
            d1 = parse_fecha(fd)
            d2 = parse_fecha(fa)
            if d1 and d2:
                tramite_str = str(abs((d2 - d1).days)) + ' d'
        aa_icon   = '<span class="ok-icon">✓</span>' if aa_ok   else '<span class="err-icon">✗</span>'
        fl_icon   = '<span class="ok-icon">✓</span>' if flujo_ok else '<span class="err-icon">✗</span>'
        table_rows += (
            f'<tr>'
            f'<td>R{code}</td>'
            f'<td>{nombre}</td>'
            f'<td class="small">{juzgado}</td>'
            f'<td class="mono">{radicado}</td>'
            f'<td>{fd}</td>'
            f'<td>{fa or "—"}</td>'
            f'<td class="center">{tramite_str}</td>'
            f'<td class="center">{aa_icon}</td>'
            f'<td class="center">{fl_icon}</td>'
            f'</tr>\n'
        )

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Dashboard NP</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#0f0f1a;color:#e0e0e0;padding:28px}}
  h2{{color:#e91e8c;margin-bottom:4px;font-size:1.3rem}}
  .sub{{color:#888;font-size:.85rem;margin-bottom:24px}}
  .kpi-row{{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:28px}}
  .kpi{{background:#1a1a2e;border-radius:12px;padding:16px 24px;min-width:120px;border:1px solid #2a2a3e}}
  .kpi .num{{font-size:2.2rem;font-weight:800;color:#e91e8c;line-height:1}}
  .kpi .lbl{{font-size:.75rem;color:#888;margin-top:6px;text-transform:uppercase;letter-spacing:.5px}}
  .kpi.ok .num{{color:#4caf50}}
  .kpi.err .num{{color:#f44336}}
  .kpi.prom .num{{color:#ff9800;font-size:1.5rem}}
  .kpi.aa .num{{color:#ab47bc}}
  .kpi.mail .num{{color:#42a5f5}}
  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:28px}}
  .chart-box{{background:#1a1a2e;border-radius:12px;padding:20px;border:1px solid #2a2a3e}}
  .chart-box h3{{color:#ccc;font-size:.93rem;margin-bottom:14px;font-weight:600}}
  .tbl-wrap{{background:#1a1a2e;border-radius:12px;padding:20px;border:1px solid #2a2a3e;overflow-x:auto;margin-bottom:28px}}
  .tbl-wrap h3{{color:#ccc;font-size:.93rem;margin-bottom:14px;font-weight:600}}
  table{{width:100%;border-collapse:collapse;font-size:.82rem}}
  thead tr{{background:#2a2a3e}}
  th{{color:#aaa;font-weight:600;padding:8px 10px;text-align:left;white-space:nowrap;border-bottom:1px solid #3a3a4e}}
  td{{padding:7px 10px;border-bottom:1px solid #1e1e30;vertical-align:middle}}
  tr:hover td{{background:#23233a}}
  .center{{text-align:center}}
  .mono{{font-family:monospace;font-size:.8rem}}
  .small{{font-size:.78rem}}
  .ok-icon{{color:#4caf50;font-size:1rem;font-weight:700}}
  .err-icon{{color:#f44336;font-size:1rem;font-weight:700}}
  .btn-row{{display:flex;gap:12px;margin-top:8px}}
  .btn{{padding:10px 24px;border-radius:8px;border:none;cursor:pointer;font-size:.9rem;font-weight:600;text-decoration:none;display:inline-flex;align-items:center;gap:6px}}
  .btn-primary{{background:#e91e8c;color:#fff}}
  .btn-primary:hover{{background:#c4006a}}
  .btn-outline{{background:transparent;color:#e91e8c;border:2px solid #e91e8c}}
  .btn-outline:hover{{background:#e91e8c;color:#fff}}
  @media(max-width:640px){{.charts{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<h2>Dashboard — Lote NP</h2>
<p class="sub">Generado el {hoy_str()}</p>

<div class="kpi-row">
  <div class="kpi">
    <div class="num">{total}</div>
    <div class="lbl">NP Procesadas</div>
  </div>
  <div class="kpi ok">
    <div class="num">{paquetes}</div>
    <div class="lbl">Exitosas</div>
  </div>
  <div class="kpi err">
    <div class="num">{errores}</div>
    <div class="lbl">Con error</div>
  </div>
  <div class="kpi aa">
    <div class="num">{aa_ok_count}</div>
    <div class="lbl">AA Extraídos</div>
  </div>
  <div class="kpi mail">
    <div class="num">{sent_ok_count}</div>
    <div class="lbl">Emails Enviados</div>
  </div>
  <div class="kpi prom">
    <div class="num">{prom_dias}</div>
    <div class="lbl">Días prom. trámite</div>
  </div>
</div>

<div class="charts">
  <div class="chart-box">
    <h3>Notificaciones Personales por Ciudad</h3>
    <canvas id="chart-ciudad" height="220"></canvas>
  </div>
  <div class="chart-box">
    <h3>Tiempo Promedio de Trámite por Ciudad (días)</h3>
    <canvas id="chart-tramite" height="220"></canvas>
  </div>
</div>

<div class="tbl-wrap">
  <h3>Detalle de Procesos</h3>
  <table>
    <thead>
      <tr>
        <th>Código</th><th>Nombre</th><th>Juzgado</th><th>Radicado</th>
        <th>Fec. Radicación</th><th>Fec. Auto Admisorio</th>
        <th>Trámite</th><th>Estado AA</th><th>Estado Flujo</th>
      </tr>
    </thead>
    <tbody>
      {table_rows}
    </tbody>
  </table>
</div>

<div class="btn-row">
  <button class="btn btn-outline" onclick="window.open('/resumen/{job_id}','_blank')">📄 Resumen</button>
  <a href="/download/{job_id}" class="btn btn-primary">⬇️ Descargar ZIP</a>
</div>

<script>
const PALETTE = ['#e91e8c','#9c27b0','#ff5722','#f06292','#673ab7','#ab47bc','#c4006a','#ff9800','#7b1fa2','#e91e63'];
function barBg(n)  {{ return Array.from({{length:n}},(_,i)=>PALETTE[i%PALETTE.length]+'cc'); }}
function barBrd(n) {{ return Array.from({{length:n}},(_,i)=>PALETTE[i%PALETTE.length]); }}

const scaleOpts = {{
  x:{{ticks:{{color:'#aaa',font:{{size:11}}}},grid:{{color:'#2a2a3e'}}}},
  y:{{ticks:{{color:'#aaa',font:{{size:11}}}},grid:{{color:'#2a2a3e'}}}}
}};

const cityLabels = {ciudad_labels_json};
new Chart(document.getElementById('chart-ciudad'), {{
  type: 'bar',
  data: {{
    labels: cityLabels,
    datasets: [{{
      data: {ciudad_values_json},
      backgroundColor: barBg(cityLabels.length),
      borderColor:     barBrd(cityLabels.length),
      borderWidth: 1, borderRadius: 6
    }}]
  }},
  options: {{plugins:{{legend:{{display:false}}}},scales:scaleOpts}}
}});

const avgLabels = {avg_labels_json};
new Chart(document.getElementById('chart-tramite'), {{
  type: 'bar',
  data: {{
    labels: avgLabels,
    datasets: [{{
      data: {avg_values_json},
      backgroundColor: barBg(avgLabels.length),
      borderColor:     barBrd(avgLabels.length),
      borderWidth: 1, borderRadius: 6
    }}]
  }},
  options: {{
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>ctx.raw+' días'}}}}}},
    scales:scaleOpts
  }}
}});
</script>
</body>
</html>"""


@app.route('/resumen/<job_id>')
def resumen(job_id):
    """Printable summary with KPIs, charts, and detail table."""
    job = JOBS.get(job_id)
    if not job:
        return "Job no encontrado", 404

    cases    = job.get('cases', [])
    paquetes = job.get('paquetes', 0)
    total    = job.get('total', 0)
    errores  = total - paquetes

    aa_ok_count   = sum(1 for c in cases if c.get('procesó_aa', False))
    sent_ok_count = sum(1 for c in cases if c.get('enviado', False))

    ciudad_counts = {}
    ciudad_days   = {}
    all_days      = []
    for c in cases:
        ciudad = c.get('Ciudad', 'N/A') or 'N/A'
        ciudad_counts[ciudad] = ciudad_counts.get(ciudad, 0) + 1
        fd = c.get('Fecha_Demanda', '')
        fa = c.get('fecha_admite_extracted', '')
        if fd and fa:
            d1 = parse_fecha(fd)
            d2 = parse_fecha(fa)
            if d1 and d2:
                days = abs((d2 - d1).days)
                all_days.append(days)
                ciudad_days.setdefault(ciudad, []).append(days)

    ciudad_labels_json = json.dumps(list(ciudad_counts.keys()))
    ciudad_values_json = json.dumps(list(ciudad_counts.values()))
    avg_cities     = list(ciudad_days.keys())
    avg_vals       = [round(sum(v)/len(v), 1) for v in ciudad_days.values()]
    avg_labels_json = json.dumps(avg_cities)
    avg_values_json = json.dumps(avg_vals)
    prom_dias       = round(sum(all_days)/len(all_days), 1) if all_days else 'N/A'
    now_str         = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # Table rows (light theme)
    table_rows = ''
    for c in cases:
        code    = c.get('code', '')
        nombre  = c.get('Nombre', '') or ''
        juzgado = c.get('Juzgado', '') or ''
        radicado= c.get('Radicado', '') or ''
        fd      = c.get('Fecha_Demanda', '') or ''
        fa      = c.get('fecha_admite_extracted', '') or ''
        aa_ok   = c.get('procesó_aa', False)
        flujo_ok= c.get('enviado', False)
        tramite_str = '—'
        if fd and fa:
            d1 = parse_fecha(fd)
            d2 = parse_fecha(fa)
            if d1 and d2:
                tramite_str = str(abs((d2 - d1).days)) + ' d'
        aa_icon  = '<span style="color:#2e7d32;font-weight:700">✓</span>' if aa_ok   else '<span style="color:#c62828;font-weight:700">✗</span>'
        fl_icon  = '<span style="color:#2e7d32;font-weight:700">✓</span>' if flujo_ok else '<span style="color:#c62828;font-weight:700">✗</span>'
        table_rows += (
            f'<tr>'
            f'<td>R{code}</td>'
            f'<td>{nombre}</td>'
            f'<td style="font-size:.78rem">{juzgado}</td>'
            f'<td style="font-family:monospace;font-size:.78rem">{radicado}</td>'
            f'<td>{fd}</td>'
            f'<td>{fa or "—"}</td>'
            f'<td style="text-align:center">{tramite_str}</td>'
            f'<td style="text-align:center">{aa_icon}</td>'
            f'<td style="text-align:center">{fl_icon}</td>'
            f'</tr>\n'
        )

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Resumen NP — QPAlliance</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#fff;color:#222;padding:32px;max-width:1000px;margin:0 auto}}
  h1{{color:#D4006A;font-size:1.6rem;margin-bottom:4px}}
  .sub{{color:#888;font-size:.85rem;margin-bottom:24px}}
  .kpi-row{{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:28px}}
  .kpi{{background:#fff5f9;border:1px solid #D4006A33;border-radius:10px;padding:14px 20px;min-width:110px}}
  .kpi .num{{font-size:1.9rem;font-weight:800;color:#D4006A}}
  .kpi .lbl{{font-size:.75rem;color:#888;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}}
  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:28px}}
  .chart-box{{border:1px solid #eee;border-radius:10px;padding:16px}}
  .chart-box h3{{color:#333;font-size:.9rem;margin-bottom:12px;font-weight:600}}
  .tbl-section{{margin-bottom:28px}}
  .tbl-section h3{{color:#333;font-size:.9rem;margin-bottom:10px;font-weight:600}}
  table{{width:100%;border-collapse:collapse;font-size:.8rem}}
  th{{background:#fce4f0;color:#9c0052;font-weight:600;padding:7px 9px;text-align:left;white-space:nowrap;border-bottom:2px solid #D4006A44}}
  td{{padding:6px 9px;border-bottom:1px solid #f0f0f0;vertical-align:middle}}
  tr:nth-child(even) td{{background:#fafafa}}
  .footer{{color:#aaa;font-size:.78rem;margin-top:32px;border-top:1px solid #eee;padding-top:12px}}
  .no-print{{margin-bottom:20px}}
  @media print{{.no-print{{display:none}};.charts{{grid-template-columns:1fr 1fr}}}}
  @media(max-width:640px){{.charts{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<div class="no-print">
  <button onclick="window.print()" style="background:#D4006A;color:#fff;border:none;padding:10px 24px;border-radius:8px;cursor:pointer;font-size:.9rem;font-weight:600">🖨️ Imprimir</button>
</div>

<h1>Resumen — Lote NP</h1>
<p class="sub">QPAlliance — Legal Department &nbsp;|&nbsp; Descargado el {now_str}</p>

<div class="kpi-row">
  <div class="kpi"><div class="num">{total}</div><div class="lbl">NP Procesadas</div></div>
  <div class="kpi"><div class="num">{paquetes}</div><div class="lbl">Exitosas</div></div>
  <div class="kpi"><div class="num">{errores}</div><div class="lbl">Con error</div></div>
  <div class="kpi"><div class="num">{aa_ok_count}</div><div class="lbl">AA Extraídos</div></div>
  <div class="kpi"><div class="num">{sent_ok_count}</div><div class="lbl">Emails Enviados</div></div>
  <div class="kpi"><div class="num">{prom_dias}</div><div class="lbl">Días prom. trámite</div></div>
</div>

<div class="charts">
  <div class="chart-box">
    <h3>NPs por Ciudad</h3>
    <canvas id="rc1" height="200"></canvas>
  </div>
  <div class="chart-box">
    <h3>Tiempo Promedio de Trámite por Ciudad (días)</h3>
    <canvas id="rc2" height="200"></canvas>
  </div>
</div>

<div class="tbl-section">
  <h3>Detalle de Procesos</h3>
  <table>
    <thead>
      <tr>
        <th>Código</th><th>Nombre</th><th>Juzgado</th><th>Radicado</th>
        <th>Fec. Radicación</th><th>Fec. Auto Admisorio</th>
        <th>Trámite</th><th>Estado AA</th><th>Estado Flujo</th>
      </tr>
    </thead>
    <tbody>
      {table_rows}
    </tbody>
  </table>
</div>

<div class="footer">Generado por Elena NP — QPAlliance — {hoy_str()}</div>

<script>
const PAL = ['#D4006A','#9c27b0','#ff5722','#f06292','#673ab7','#ab47bc','#c4006a','#ff9800','#7b1fa2','#e91e63'];
function bg(n)  {{ return Array.from({{length:n}},(_,i)=>PAL[i%PAL.length]+'bb'); }}
function brd(n) {{ return Array.from({{length:n}},(_,i)=>PAL[i%PAL.length]); }}
const sc = {{x:{{ticks:{{font:{{size:10}}}}}},y:{{ticks:{{font:{{size:10}}}}}}}};
const cityL = {ciudad_labels_json};
new Chart(document.getElementById('rc1'),{{
  type:'bar',
  data:{{labels:cityL,datasets:[{{data:{ciudad_values_json},backgroundColor:bg(cityL.length),borderColor:brd(cityL.length),borderWidth:1,borderRadius:4}}]}},
  options:{{plugins:{{legend:{{display:false}}}},scales:sc}}
}});
const avgL = {avg_labels_json};
new Chart(document.getElementById('rc2'),{{
  type:'bar',
  data:{{labels:avgL,datasets:[{{data:{avg_values_json},backgroundColor:bg(avgL.length),borderColor:brd(avgL.length),borderWidth:1,borderRadius:4}}]}},
  options:{{plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>ctx.raw+' días'}}}}}},scales:sc}}
}});
</script>
</body>
</html>"""

# ─── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
